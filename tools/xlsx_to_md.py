#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""แปลงไฟล์ .xlsx เป็น Markdown ทุกชีต (ตารางเต็ม ไม่ตัดทอน)

    python3 tools/xlsx_to_md.py "<in.xlsx>" "<out.md>"

ใช้กับ SBP/TSM-SRM-LLDD SBP workflow 1.2.xlsx เพื่อให้ทีมอ่านนิยาม workflow engine
ได้จาก markdown โดยไม่ต้องเปิด Excel · ไฟล์ต้นทางใน SBP/ เป็น read-only (ไม่แก้ไข)
"""
from __future__ import annotations
import sys, pathlib, datetime
import openpyxl


def cell(v) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime.datetime):
        return v.strftime("%d/%m/%Y") if (v.hour, v.minute, v.second) == (0, 0, 0) else v.strftime("%d/%m/%Y %H:%M")
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    s = str(v).replace("\r\n", "\n").strip()
    # ตัวคั่นตารางกับ newline ต้อง escape ไม่งั้นตาราง markdown พัง
    return s.replace("|", "\\|").replace("\n", "<br>")


def sheet_to_md(ws) -> list[str]:
    rows = [[cell(c) for c in row] for row in ws.iter_rows(values_only=True)]
    while rows and not any(rows[-1]):
        rows.pop()
    if not rows:
        return ["_(ชีตว่าง)_", ""]
    width = max((len(r) for r in rows), default=0)
    while width and all(len(r) <= width - 1 or not r[width - 1] for r in rows):
        width -= 1
    if width == 0:
        return ["_(ชีตว่าง)_", ""]
    rows = [(r + [""] * width)[:width] for r in rows]

    # แถวแรกที่มีข้อมูล >= 2 ช่อง ถือเป็นหัวตาราง ที่เหลือเป็นเนื้อ
    head_i = next((i for i, r in enumerate(rows) if sum(1 for x in r if x) >= 2), 0)
    out: list[str] = []
    for r in rows[:head_i]:
        text = " ".join(x for x in r if x)
        if text:
            out += [f"**{text}**", ""]
    head = rows[head_i]
    head = [h or f"col{i+1}" for i, h in enumerate(head)]
    out.append("| " + " | ".join(head) + " |")
    out.append("|" + "|".join(["---"] * width) + "|")
    for r in rows[head_i + 1:]:
        if not any(r):
            continue
        out.append("| " + " | ".join(r) + " |")
    out.append("")
    return out


def main() -> None:
    if len(sys.argv) != 3:
        raise SystemExit(__doc__)
    src, dst = pathlib.Path(sys.argv[1]), pathlib.Path(sys.argv[2])
    wb = openpyxl.load_workbook(src, data_only=True)
    lines = [
        f"# {src.stem} — แปลงจาก Excel เป็น Markdown",
        "",
        f"> แปลงอัตโนมัติจาก `{src}` ({src.stat().st_size // 1024:,} KB · {len(wb.sheetnames)} ชีต) "
        f"ด้วย `tools/xlsx_to_md.py` — **ไฟล์ต้นทางเป็น read-only ไม่ถูกแก้ไข**",
        "> ต้องการเนื้อหาสรุปอ่านง่ายให้ดู [`SBP/TSM-SRM-LLDD-SBP-workflow-1.2.md`](../SBP/TSM-SRM-LLDD-SBP-workflow-1.2.md) แทน",
        "",
        "## สารบัญชีต",
        "",
        "| # | ชีต | ขนาด |",
        "|---|---|---|",
    ]
    for i, ws in enumerate(wb.worksheets, 1):
        anchor = ws.title.lower().replace(" ", "-")
        lines.append(f"| {i} | [{ws.title}](#{anchor}) | {ws.max_row} แถว × {ws.max_column} คอลัมน์ |")
    lines.append("")
    for ws in wb.worksheets:
        lines += ["---", "", f"## {ws.title}", ""]
        lines += sheet_to_md(ws)
    dst.parent.mkdir(parents=True, exist_ok=True)
    dst.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"{dst} · {len(wb.sheetnames)} ชีต · {dst.stat().st_size // 1024:,} KB")


if __name__ == "__main__":
    main()
